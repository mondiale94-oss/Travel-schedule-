#!/usr/bin/env python3
"""
Generate travel-availability.xlsx with sheets Input, Dates, Summary, Help.
Usage:
  pip install openpyxl
  python scripts/generate_travel_availability.py

This script creates travel-availability.xlsx in the repository root.
"""
from openpyxl import Workbook
from datetime import date, timedelta


def create_workbook(filename='travel-availability.xlsx', days=30):
    wb = Workbook()

    # --- Input sheet ---
    ws_in = wb.active
    ws_in.title = "Input"
    headers = ["الاسم", "دولة A من", "دولة A إلى", "دولة B من", "دولة B إلى"]
    ws_in.append(headers)

    # Example rows (you can remove or edit these in the generated file)
    ws_in.append(["أحمد", date.today(), date.today() + timedelta(days=5), None, None])
    ws_in.append(["منى", None, None, date.today(), date.today() + timedelta(days=10)])

    # --- Dates sheet ---
    ws_dates = wb.create_sheet("Dates")
    ws_dates.append(["التاريخ", "متاح في دولة A", "متاح في دولة B"]) 

    start = date.today()
    for i in range(days):
        d = start + timedelta(days=i)
        ws_dates.append([d])

    # Add COUNTIFS formulas for each date row (starts at row 2)
    for idx in range(2, 2 + days):
        date_cell = f"A{idx}"
        # COUNTIFS(Input!$B:$B,"<="&A2, Input!$C:$C,">="&A2)
        countA = f"=COUNTIFS(Input!$B:$B,\"<=\"&{date_cell},Input!$C:$C,\">=\"&{date_cell})"
        countB = f"=COUNTIFS(Input!$D:$D,\"<=\"&{date_cell},Input!$E:$E,\">=\"&{date_cell})"
        ws_dates[f"B{idx}"] = countA
        ws_dates[f"C{idx}"] = countB

    # --- Summary sheet ---
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["ملخّص التوافر"]) 
    ws_sum.append(["أفضل عدد متاح لدولة A", "=MAX(Dates!B:B)"])
    ws_sum.append(["أفضل عدد متاح لدولة B", "=MAX(Dates!C:C)"])
    ws_sum.append(["أول تاريخ بأفضل توافر A", "=INDEX(Dates!$A:$A,MATCH(MAX(Dates!B:B),Dates!$B:$B,0))"]) 
    ws_sum.append(["أول تاريخ بأفضل توافر B", "=INDEX(Dates!$A:$A,MATCH(MAX(Dates!C:C),Dates!$C:$C,0))"]) 

    # --- Help sheet ---
    ws_help = wb.create_sheet("Help")
    help_text = [
        "شرح: هذا الملف يُنشأ تلقائياً بواسطة السكربت.",
        "ورقة Input: أدخل كل شخص ونطاق التواريخ المتاحة لكل دولة (التواريخ كـ YYYY-MM-DD أو كتواريخ Excel).",
        "أعمدة Input هي: الاسم | دولة A من | دولة A إلى | دولة B من | دولة B إلى",
        "ورقة Dates: تولّد تواريخ (افتراضياً 30 يوم بدءاً من اليوم) وتحسب عدد الأشخاص المتاحين لكل تاريخ لكل دولة.",
        "ورقة Summary: تعرض أفضل عدد من الأشخاص المتاحين لكل دولة وتعرض أول تاريخ يصل للحد الأعظمي.",
        "لتغيير طول النافذة الزمنية عدّل قيمة days (الافتراضية 30) عند استدعاء create_workbook أو سطّر السكربت.",
        "لتشغيل السكربت: pip install openpyxl && python scripts/generate_travel_availability.py",
        "بعد التشغيل سيتم حفظ travel-availability.xlsx في جذر المستودع - يمكنك رفعه أو فتحه مباشرة.",
    ]
    for i, t in enumerate(help_text, start=1):
        ws_help[f"A{i}"] = t

    wb.save(filename)
    print(f"Saved {filename}")


if __name__ == '__main__':
    create_workbook()
